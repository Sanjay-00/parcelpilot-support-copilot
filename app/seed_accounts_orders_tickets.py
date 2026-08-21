import sqlite3

import openpyxl


def _to_int_or_none(value):
    if value is None:
        return None
    return 1 if value else 0


def load(conn: sqlite3.Connection, xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    accounts_ws = wb["accounts"]
    rows = list(accounts_ws.iter_rows(values_only=True))
    for account_id, account_name, plan, status, csm, contract_file, premium_support, notes in rows[1:]:
        conn.execute(
            "INSERT INTO accounts (account_id, account_name, plan, status, csm, "
            "contract_file, premium_support, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (account_id, account_name, plan, status, csm, contract_file,
             _to_int_or_none(premium_support), notes),
        )

    orders_ws = wb["orders"]
    rows = list(orders_ws.iter_rows(values_only=True))
    for (order_id, account_id, carrier, status, booked_at, pickup_window_start,
         pickup_window_end, pickup_actual_at, shipment_fee_inr, carrier_fault,
         customer_fault, cancellation_requested_at, notes) in rows[1:]:
        conn.execute(
            "INSERT INTO orders (order_id, account_id, carrier, status, booked_at, "
            "pickup_window_start, pickup_window_end, pickup_actual_at, shipment_fee_inr, "
            "carrier_fault, customer_fault, cancellation_requested_at, notes) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (order_id, account_id, carrier, status, str(booked_at),
             str(pickup_window_start) if pickup_window_start else None,
             str(pickup_window_end) if pickup_window_end else None,
             str(pickup_actual_at) if pickup_actual_at else None,
             shipment_fee_inr, _to_int_or_none(carrier_fault),
             _to_int_or_none(customer_fault),
             str(cancellation_requested_at) if cancellation_requested_at else None,
             notes),
        )

    tickets_ws = wb["tickets"]
    rows = list(tickets_ws.iter_rows(values_only=True))
    for (ticket_id, account_id, created_at, status, subject, description, channel,
         assigned_to, last_customer_message_at, historical_resolution) in rows[1:]:
        conn.execute(
            "INSERT INTO tickets (ticket_id, account_id, created_at, status, subject, "
            "description, channel, assigned_to, last_customer_message_at, "
            "historical_resolution) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (ticket_id, account_id, str(created_at), status, subject, description,
             channel, assigned_to,
             str(last_customer_message_at) if last_customer_message_at else None,
             historical_resolution),
        )

    conn.commit()
